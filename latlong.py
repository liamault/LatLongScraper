from tkinter import filedialog
import openpyxl as opx
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from time import sleep
import cv2
import numpy as np
import os


def getPingCoord(imgpath):
    im = cv2.imread(imgpath)
    red = [17,21,178]
    #color changes when run with head, which breaks function
    yarr, xarr = np.where(np.all(im==red,axis=2))

    return {"x":xarr[34]-100,"y":yarr[34]+23}

#creates browser
options = Options()
options.headless = True
options.add_argument("--log-level=3")
options.add_argument("--window-size=1920,1080")
driver = webdriver.Chrome(options=options)
driver.get('https://www.google.com/maps')
tb = driver.find_element(By.ID,"searchboxinput")
ac = ActionChains(driver)

fpath = filedialog.askopenfilename()

wb = opx.load_workbook(filename=fpath)
ws = wb.worksheets[0]

#change 1 to a 2 if first row of excel is labels
for r in range(1,ws.max_row+1):
    #adjust
    addr = str(ws['A' + str(r)].value) + ", " + str(ws['B' + str(r)].value) + ", " + str(ws['C' + str(r)].value)
    print(addr)

    tb.clear()
    tb.send_keys(addr)
    tb.send_keys(webdriver.common.keys.Keys.ENTER)
    
    #may have to adjust this depending on machine/internet
    sleep(2)

    bt = driver.find_element(By.CSS_SELECTOR,"#widget-zoom-in")
    bt.click()
    bt.click()
    bt.click()
    bt.click()

    driver.save_screenshot("ss.png")
    pingCoord = getPingCoord("ss.png")
        
    ac.move_to_element(tb).move_by_offset((pingCoord['x']-tb.location['x']),(pingCoord['y']-tb.location['y'])).context_click().perform()
    #may have to adjust this depending on machine/internet
    sleep(0.5)
    adCoord = driver.find_element(By.CLASS_NAME,"mLuXec")
    print(adCoord.text)

    #writes lat long to excel, columns can be changed by changing the letter
    ws['F' + str(r)].value = float(adCoord.text.split(',')[0])
    ws['G' + str(r)].value = float(adCoord.text.split(' ')[1])

driver.close()
os.remove("ss.png")
wb.save(filename=fpath)
wb.close()
